from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

class ExcelUtil:

    def __init__(self):

        self.headerBorder = Border(
            left = Side(Border(border_style='thin', color='FF000000')),
            right = Side(Border(border_style='thin', color='FF000000')),
            top = Side(Border(border_style='thin', color='FF000000')),
            bottom = Side(Border(border_style='thin', color='FF000000')),
            diagonal = Side(Border(border_style='thin', color='FF000000')),
            diagonal_direction = 0,
            outline = Side(Border(border_style='thin', color='FF000000')),
            vertical = Side(Border(border_style='thin', color='FF000000')),
            horizontal = Side(Border(border_style='thin', color='FF000000'))
        )

        self.headerBgcolor = PatternFill(
            fill_type='solid',
            fgColor=Color('FFC000')
        )

        self.headerFont.Font = Font(bold=True)

        self.dataBorder = Border(
            left = Side(Border(border_style='thin', color='FF000000')),
            right = Side(Border(border_style='thin', color='FF000000')),
            top = Side(Border(border_style='thin', color='FF000000')),
            bottom = Side(Border(border_style='thin', color='FF000000')),
            diagonal = Side(Border(border_style='thin', color='FF000000')),
            diagonal_direction = 0,
            outline = Side(Border(border_style='thin', color='FF000000')),
            vertical = Side(Border(border_style='thin', color='FF000000')),
            horizontal = Side(Border(border_style='thin', color='FF000000'))
        )

        self.dataBgcolor = PatternFill(
            fill_type='solid',
            fgColor=Color('000000')
        )
        
    def createFile(self, filename, data, headerData = None, isFirstHeaderRow = False):

        workbook = Workbook()
        worksheet = workbook.active
        startIdx = 0
        excelStartIdx = 0

        if headerData != None:

            for x in range(1, len(headerData) + 1):
                worksheet.cell(row=1, column=x).value = headerData[x-1]
                worksheet.cell(row=1, column=x).border = self.headerBorder
                worksheet.cell(row=1, column=x).fill = self.headerBgcolor
                worksheet.cell(row=1, column=x).font = self.headerFont

            excelStartIdx += 1

            if isFirstHeaderRow == True:

                for x in range(1, len(data[0] + 1)):
                    worksheet.cell(row=1, column=x).value = data[0][x-1]
                    worksheet.cell(row=1, column=x).border = self.headerBorder
                    worksheet.cell(row=1, column=x).fill = self.headerBgcolor
                    worksheet.cell(row=1, column=x).font = self.headerFont

                startIdx += 1

        for y in range(1 + startIdx, len(data) + 1):
            for x in range(1, len(data[0]) + 1):
                worksheet.cell(row = (y + excelStartIdx), column = x).value = data[y - 1][x - 1]
                worksheet.cell(row = (y + excelStartIdx), column = x).border = self.dataBorder
                worksheet.cell(row = (y + excelStartIdx), column = x).fill = self.dataBgcolor

        workbook.save(filename)
